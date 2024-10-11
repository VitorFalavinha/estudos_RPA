from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui as pyto
import re
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime

def get_data_from_url():
    # Configuração do driver do Selenium
    chrome_options = Options()
    chrome_options.headless = True  # Mudar para True se quiser rodar sem interface gráfica

    service = Service('C:/Users/vitor/OneDrive/Documentos/EstudosRPA/estudos_RPA/chromedriver.exe') 

    # Inicializar o driver do Selenium
    driver = webdriver.Chrome(service=service, options=chrome_options)

    # URL do vídeo do YouTube
    url = 'https://www.youtube.com/watch?v=e4Yt-UlsJG4' 

     
    # Abrir a página do vídeo
    driver.get(url)
    driver.maximize_window()

    pyto.PAUSE = 7

    # Obtém as dimensões da tela
    screen_width, screen_height = pyto.size()

    # Calcula o ponto central da tela
    center_x = screen_width / 2
    center_y = screen_height / 2

    # Move o mouse para o centro da tela
    pyto.moveTo(center_x, center_y)

    #scroll down para visualizar comentários
    pyto.vscroll(-500) 
    
    # Clique no centro da tela para capturar as visualizações detalhadas
    pyto.click()

    try:
        # Espera explícita para o elemento estar presente e visível
        views_element = WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located((By.XPATH, "/html/body/ytd-app/div[1]/ytd-page-manager/ytd-watch-flexy/div[5]/div[1]/div/div[2]/ytd-watch-metadata/div/div[4]/div[1]/div/ytd-watch-info-text/div/yt-formatted-string/span[1]"))
        )
    
        # Obter o texto das visualizações e comentários
        views = views_element.text
        print(f"O vídeo tem {views}.")

    except Exception as e:
        print(f"Erro ao encontrar o elemento: {e}")


    try:
        # Espera explícita para o elemento estar presente e visível
        comments_element = WebDriverWait(driver, 40).until(
            EC.visibility_of_element_located((By.XPATH, '/html/body/ytd-app/div[1]/ytd-page-manager/ytd-watch-flexy/div[5]/div[1]/div/div[2]/ytd-comments/ytd-item-section-renderer/div[1]/ytd-comments-header-renderer/div[1]/div[1]'))
        )
    
        # Obter o texto das visualizações e comentários
        comments = comments_element.text
        print(f"O vídeo tem {comments}.")

        # Extraindo apenas os números usando expressão regular
        views_number = int(re.sub(r'[^\d]', '', views_element.text))
        comments_number = int(re.sub(r'[^\d]', '', comments_element.text))
        
        print(f"Número de visualizações: {views_number}")
        print(f"Número de comentários: {comments_number}")

        # Obtendo a data e hora atuais
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        fill_in_excel(url, views_number, comments_number, current_datetime )

    except Exception as e:
        print(f"Erro ao encontrar o elemento: {e}")
        
    finally:
        # Fechar o navegador
        driver.quit()
    
    return (views_number, comments_number)

def fill_in_excel(url, views_number, comments_number, current_datetime):
    # Criação e preenchimento da planilha Excel
    #wb = Workbook()
    #ws = wb.active
    #ws.title = "YouTube_Data"

    # Carregando a planilha existente
    wb = load_workbook("C:/Users/vitor/OneDrive/Documentos/EstudosRPA/estudos_RPA/data.xlsm", keep_vba=True)
    ws = wb['YouTube_Data']  

    # Escrevendo os cabeçalhos
    #ws.append(["URL", "Visualizações", "Comentários","Data da atualização"])

    # Adicionando uma nova linha com os dados
    ws.append([url, views_number, comments_number, current_datetime])
    
    # Salvando a planilha com os novos dados
    wb.save("C:/Users/vitor/OneDrive/Documentos/EstudosRPA/estudos_RPA/data.xlsm")

    update_excel()

def update_excel():
    pyto.PAUSE = 5
    pyto.press('Win')
    pyto.write('estudos_RPA')
    pyto.press('Enter')
    pyto.press('d')
    pyto.press('Enter')
    pyto.hotkey('Alt', 'F4')
    pyto.press('Enter')


get_data_from_url()